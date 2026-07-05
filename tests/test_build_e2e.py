"""E2E 测试：build.py 用 test.xlsx 生成 output/index.html，验证结构。"""
import subprocess
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent


def test_build_runs():
    """build.py 能跑通并生成 output/index.html。"""
    # 确保 test.xlsx 存在
    if not (PROJECT_ROOT / "data" / "test.xlsx").exists():
        subprocess.run(
            [sys.executable, str(PROJECT_ROOT / "tests" / "_make_test_xlsx.py"), "data/test.xlsx"],
            check=True, cwd=PROJECT_ROOT,
        )

    result = subprocess.run(
        [sys.executable, "build.py"],
        capture_output=True, text=True, cwd=PROJECT_ROOT,
    )
    assert result.returncode == 0, f"build failed: {result.stderr}"
    assert "Built" in result.stdout

    output = PROJECT_ROOT / "output" / "index.html"
    assert output.exists()
    return output


def test_output_has_6_modules():
    """输出含 6 个模块。"""
    output = PROJECT_ROOT / "output" / "index.html"
    if not output.exists():
        test_build_runs()
    html = output.read_text(encoding="utf-8")
    for title in ["一、组织架构", "二、员工基本情况", "三、人员优化情况", "四、干部队伍建设", "五、考核薪酬", "六、培训赋能"]:
        assert title in html, f"missing module: {title}"


def test_output_has_no_external_deps():
    """输出无 CDN / 外部 fetch。"""
    html = (PROJECT_ROOT / "output" / "index.html").read_text(encoding="utf-8")
    for bad in ["cdn.jsdelivr.net", "googleapis.com", "unpkg.com"]:
        assert bad not in html, f"found external dep: {bad}"


def test_output_size_under_110kb():
    """文件 ≤ 110KB（v1.5.21 + v1.5.21.2 + v1.5.21.3 mobile 适配后实际 ~102KB）。"""
    size = (PROJECT_ROOT / "output" / "index.html").stat().st_size
    assert size <= 110 * 1024, f"output is {size} bytes, exceeds 110KB budget"


def test_m4_uses_tree_renders():
    """M-4 干部队伍建设 用 v2 分类树 渲染（v1.5.13+）。"""
    html = (PROJECT_ROOT / "output" / "index.html").read_text(encoding="utf-8")
    assert "干部队伍建设总览" in html
    assert "市管领导干部" in html
    assert "高管人员" in html
    assert "派驻纪检组" in html
    assert "中层干部" in html
    assert 'tn root' in html
    assert 'branch core' in html


def test_data_composition_uses_donut():
    """组成关系用 donut 表达（M-1 营业状态等仍用 donut）。"""
    html = (PROJECT_ROOT / "output" / "index.html").read_text(encoding="utf-8")
    assert html.count("donut-svg") >= 1, "至少 M-1 营业状态等用 donut 表达"


# ====== v1.4 E2E ======

def test_delta_auto_calc():
    """v1.4 T16: 二次构建后 _prev sheet 包含上期数据 + delta 字段非空。"""
    import openpyxl
    from composer import calc_delta

    # 确保 _prev 已写（build 一次后再跑一次）
    subprocess.run([sys.executable, "build.py"], capture_output=True, cwd=PROJECT_ROOT)
    wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "test.xlsx", data_only=True)
    assert "_prev" in wb.sheetnames, "_prev sheet 应在二次构建后存在"
    prev_rows = list(wb["_prev"].iter_rows(values_only=True))
    assert len(prev_rows) >= 2, "_prev sheet 应有表头 + 至少 1 条数据"
    wb.close()

    # calc_delta 数值正确
    assert calc_delta(100, 80) == {"abs": 20, "pct": 0.25, "direction": "up", "display": "+25.0%"}
    assert calc_delta(50, None) is None
    assert calc_delta(50, 0) is None


def test_sub_text_render():
    """v1.4 T17: 含 sub_text 的模块 HTML 含 narrative-grid；metric_note 渲染到 KPI 卡内。"""
    output = PROJECT_ROOT / "output" / "index.html"
    subprocess.run([sys.executable, "build.py"], capture_output=True, cwd=PROJECT_ROOT)
    html = output.read_text(encoding="utf-8")
    assert html.count("narrative-grid") >= 1, "narrative-grid 应至少出现 1 次（子包含 sub_text 的模块）"
    assert html.count("kpi-note") >= 1, "kpi-note 应至少出现 1 次"


def test_3color_risk_level():
    """v1.4 T18: 含完成率数据的 KPI 卡有 data-category 属性；语义正确。"""
    from composer import classify_rate

    # 语义测试
    assert classify_rate(105) == "success"
    assert classify_rate(95) == "warning"
    assert classify_rate(50) == "danger"

    # HTML 测试
    subprocess.run([sys.executable, "build.py"], capture_output=True, cwd=PROJECT_ROOT)
    html = (PROJECT_ROOT / "output" / "index.html").read_text(encoding="utf-8")
    assert html.count("data-category=") >= 1, "应有 KPI 卡含 data-category 属性"
    # 至少有一个 success/warning/danger 类别
    assert any(cat in html for cat in ['data-category="success"', 'data-category="warning"', 'data-category="danger"'])


def test_m2_heatmap():
    """v1.4 T19: render_heatmap 函数可用 + 输出标准结构。

    注：当前 M-2 数据不是二维"员工 × 司龄"结构（heat不触发），
    但 viz/heatmap.py 函数需存在并输出 25 cells。
    """
    from viz.heatmap import render_heatmap
    cells = []
    for r in ["R1","R2","R3","R4","R5"]:
        for c in ["C1","C2","C3","C4","C5"]:
            cells.append({"row": r, "col": c, "value": 0.5, "level": "success"})
    html = render_heatmap(cells, {})
    assert '<div class="heatmap">' in html
    assert html.count("heatmap-cell") == 25
    assert html.count("data-level") == 25
