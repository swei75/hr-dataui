"""生成测试 Excel - 完整 .docx 数据。

每行格式: [section, label, value, sub]
section → 对应 mapping.py 的 data_key
"""
import sys
from pathlib import Path
import openpyxl


def make_test_xlsx(out_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 配置
    ws = wb.create_sheet("配置")
    ws.append(["key", "value"])
    ws.append(["报告标题", "人力资源管理数据驾驶舱"])
    ws.append(["报告期", "2026年5月"])
    ws.append(["机构名", "XX 银行"])
    ws.append(["主色", "#1e5baa"])

    # ============ M-1 组织架构 ============
    ws = wb.create_sheet("组织架构")
    ws.append(["section", "label", "value", "sub"])
    # 总行架构
    ws.append(["headquarters", "总行本部一级部门", 28, "个"])
    ws.append(["headquarters", "总行直属一级部门", 2, "个"])
    ws.append(["headquarters", "总行直属二级部门", 13, "个"])
    # 分行
    ws.append(["branches", "分行级机构", 24, "家（含总行营业部、信用卡中心）"])
    ws.append(["branches", "全行网点", 183, "家"])
    ws.append(["branches", "正常营业", 179, "家"])
    ws.append(["branches", "临时停业", 4, "家"])
    # 广州 107 家
    for name, count in [
        ("行总营业部", 1), ("广州分行", 37), ("海珠中心支行", 22),
        ("天河中心支行", 14), ("白云中心支行", 12), ("开发区中心支行", 11),
        ("增城中心支行", 4), ("南沙分行", 5), ("科技支行", 1),
    ]:
        ws.append(["branch_guangzhou", name, count, "家"])
    # 异地 76 家
    for name, count in [
        ("深圳分行", 12), ("佛山分行", 12), ("南京分行", 9), ("惠州分行", 7),
        ("江门分行", 6), ("中山分行", 8), ("肇庆分行", 4), ("东莞分行", 8),
        ("横琴分行", 4), ("清远分行", 2), ("汕头分行", 1), ("韶关分行", 1),
        ("湛江分行", 1), ("梅州分行", 1),
    ]:
        ws.append(["branch_other", name, count, "家"])

    # ============ M-2 员工基本情况 ============
    ws = wb.create_sheet("员工情况")
    ws.append(["section", "label", "value", "sub"])
    # 总量
    ws.append(["headcount", "正式员工", 5901, "人"])
    ws.append(["headcount", "劳务派遣", 688, "人"])
    ws.append(["headcount", "合计总数", 6589, "人"])
    ws.append(["headcount", "总行机关", 716, "12.13%"])
    ws.append(["headcount", "总行直属机构", 476, "8.07%"])
    ws.append(["headcount", "经营机构", 4709, "79.80%"])
    ws.append(["headcount", "前台", 3254, "人"])
    ws.append(["headcount", "中后台", 2372, "人"])
    ws.append(["headcount", "前中后台比", "1:0.73", ""])
    ws.append(["headcount", "平均年龄", 36, "岁"])
    # 年龄分布
    for name, val, pct in [("30岁以下", 1526, 0.2586), ("31-40岁", 2851, 0.4831), ("41-50岁", 1123, 0.1903), ("51岁以上", 401, 0.068)]:
        ws.append(["age_dist", name, val, f"{pct*100:.2f}%"])
    # 性别
    for name, val, pct in [("男", 2717, 0.4604), ("女", 3184, 0.5396)]:
        ws.append(["gender_dist", name, val, f"{pct*100:.2f}%"])
    # 学历
    for name, val, pct in [("硕士研究生及以上", 1031, 0.1747), ("本科", 4514, 0.765), ("专科", 326, 0.0552), ("专科以下", 30, 0.0051)]:
        ws.append(["edu_dist", name, val, f"{pct*100:.2f}%"])
    # 客户经理
    ws.append(["customer_managers", "客户经理合计", 1692, "不含 172 劳务派遣"])
    for name, val in [("对公客户经理", 850), ("个人客户经理", 390), ("个贷客户经理", 359), ("微贷客户经理", 93)]:
        ws.append(["customer_managers", name, val, "人"])
    # 客服经理
    ws.append(["service_managers", "客服经理合计", 1165, "人"])
    for name, val in [("客服经理（非临柜）", 163), ("客服经理（临柜）", 1002)]:
        ws.append(["service_managers", name, val, "人"])

    # ============ M-3 人员优化 ============
    ws = wb.create_sheet("人员优化")
    ws.append(["section", "label", "value", "sub"])
    # 社招
    ws.append(["social_hire", "社招合计", 34, "人 · 2026 年以来"])
    for name, val in [
        ("总行总经理助理级以上", 2), ("分支行部门班子", 4), ("网点班子", 4),
        ("各类客户经理", 9), ("其他", 15),
    ]:
        ws.append(["social_hire", name, val, "人"])
    # 校招
    ws.append(["campus_hire", "全年目标", 659, "人"])
    ws.append(["campus_hire", "已签约/签约中", 721, "人"])
    ws.append(["campus_hire", "秋招", 250, "人"])
    ws.append(["campus_hire", "春招", 430, "人"])
    ws.append(["campus_hire", "微贷/特资", 41, "人"])
    ws.append(["campus_hire", "一本及以上占比", "73%", "+7% YoY"])
    ws.append(["campus_hire", "总行管培生", 45, "人"])
    ws.append(["campus_hire", "总行金融科技岗", 54, "人"])
    ws.append(["campus_hire", "分行管培生", 112, "人"])
    ws.append(["campus_hire", "英才岗", 469, "人"])
    ws.append(["campus_hire", "微贷/特资定向", 41, "人"])
    # 退出
    ws.append(["exit", "累计退出", 202, "人 · 2026 年以来"])
    ws.append(["exit", "总行直属 离职", 24, "人"])
    ws.append(["exit", "总行直属 退休", 6, "人"])
    ws.append(["exit", "经营机构 离职", 153, "人"])
    ws.append(["exit", "经营机构 退休", 19, "人"])
    ws.append(["exit", "高管级退休", 1, "人（胡优华资深督导）"])
    ws.append(["exit", "中层干部 离职", 2, "人"])
    ws.append(["exit", "员工 离职", 175, "人"])
    ws.append(["exit", "员工 退休", 24, "人"])
    # 腾笼换鸟 Top3 / Bottom3
    ws.append(["teng_long", "全年任务", 410, "人"])
    ws.append(["teng_long", "1-5月累计退出", 201, "人"])
    ws.append(["teng_long", "任务完成率", "50%", ""])
    ws.append(["teng_long", "完成率 Top1", "总营 295%", "序时"])
    ws.append(["teng_long", "完成率 Top2", "惠州 259%", "序时"])
    ws.append(["teng_long", "完成率 Top3", "东莞 255%", "序时"])
    ws.append(["teng_long", "完成量 Top1", "广分 28人", "人"])
    ws.append(["teng_long", "完成量 Top2", "深圳 23人", "人"])
    ws.append(["teng_long", "完成量 Top3", "海珠 20人", "人"])
    ws.append(["teng_long", "完成率 Bottom1", "佛山 44%", "序时"])
    ws.append(["teng_long", "完成率 Bottom2", "汕头 48%", "序时"])
    ws.append(["teng_long", "完成率 Bottom3", "广分 80%", "序时"])
    # 绩差退出
    ws.append(["poor_perf", "目标退出率", "50%", "绩差客户经理"])
    ws.append(["poor_perf", "1-5月累计退出", 44, "人"])
    ws.append(["poor_perf", "完成率", "46%", "占绩差总人数"])
    ws.append(["poor_perf", "退出率 Top1-5", "开发区/南沙/江门/汕头/梅州 100%", "并列"])
    ws.append(["poor_perf", "退出人数 Top1", "深圳 8人", ""])
    ws.append(["poor_perf", "退出人数 Top2", "东莞 7人", ""])
    ws.append(["poor_perf", "退出人数 Top3", "海珠 6人", ""])
    ws.append(["poor_perf", "退出率 Bottom1", "南京 20%", ""])
    ws.append(["poor_perf", "退出率 Bottom2", "中山 20%", ""])
    ws.append(["poor_perf", "退出率 Bottom3", "深圳 27%", ""])
    # 劳动仲裁
    ws.append(["labor", "1-5月发生", 15, "人"])
    ws.append(["labor", "已结案", 4, "人"])
    ws.append(["labor", "进行中", 11, "人"])
    ws.append(["labor", "胜诉", 2, "人（广州分行、信用卡中心）"])
    ws.append(["labor", "败诉", 2, "人（汕头分行、信用卡中心）"])

    # ============ M-4 干部队伍 ============
    ws = wb.create_sheet("干部队伍")
    ws.append(["section", "label", "value", "sub"])
    ws.append(["overview", "中层干部总职数", 211, "人"])
    ws.append(["overview", "总行部门", 100, "人"])
    ws.append(["overview", "经营机构", 111, "人"])
    ws.append(["overview", "中层正职", 54, "人"])
    ws.append(["overview", "中层副职", 157, "人"])
    ws.append(["overview", "已配备", 194, "人（92%）"])
    ws.append(["overview", "尚空缺", 17, "人（8%）"])
    ws.append(["overview", "总行已配备", 93, "人（空缺 7）"])
    ws.append(["overview", "经营已配备", 101, "人（空缺 10）"])
    ws.append(["overview", "正职已配备", 50, "人（空缺 4）"])
    ws.append(["overview", "副职已配备", 144, "人（空缺 13）"])
    # 职数表
    for row in [
        ("position_table", "市管领导干部", 8, 8, 0, ""),
        ("position_table", "高管人员", None, 6, None, "含工会主席、行长助理、首席风险官、首席信息官"),
        ("position_table", "派驻纪检监察组（不含组长）", 6, 3, 3, ""),
        ("position_table", "中层正职", 54, 50, 4, "副总级一把手 17 名"),
        ("position_table", "中层副职", 157, 144, 13, "副总 67 + 总助 74 + 经理 3，含 3 纪检"),
        ("position_table", "总行部门班子", 100, 93, 7, ""),
        ("position_table", "经营机构班子", 111, 101, 10, "含 3 纪检"),
    ]:
        ws.append(list(row))
    # 调整
    ws.append(["adjustments", "提拔", 94, "人次"])
    ws.append(["adjustments", "降职、免职（含受处分）", 14, "人次"])
    ws.append(["adjustments", "到龄转岗", 13, "人次"])

    # ============ M-5 考核薪酬 ============
    ws = wb.create_sheet("考核薪酬")
    ws.append(["section", "label", "value", "sub"])
    # 薪酬
    ws.append(["salary", "2025 年工资总额", "15", "亿元"])
    ws.append(["salary", "2025 年人均薪酬", "24.3", "万元（含预发 2026 开门红 0.3 亿）"])
    ws.append(["salary", "2026 年 1-5 月已发放", "4.2", "亿元"])
    # 年金 + 医疗
    ws.append(["pension", "2025 年企业年金余额", "16.37", "亿元"])
    ws.append(["pension", "2026.5 余额", "17.44", "亿元"])
    ws.append(["pension", "个人缴费比例", "2%", ""])
    ws.append(["pension", "单位缴费比例", "8%", ""])
    ws.append(["pension", "2025 补充医疗缴费", "7979", "万元"])
    ws.append(["pension", "节余总额", "2", "亿元（含个人 + 单位住院）"])
    ws.append(["pension", "2026.5 节余", "1.9", "亿元（个人 1.6 + 单位 3002 万）"])
    # 考核
    ws.append(["monitoring", "监测人数", 95, "经营机构班子成员（不含到龄转岗及离职脱密期）"])
    ws.append(["monitoring", "监测人次", 113, "人 · 2026 Q1"])
    ws.append(["monitoring", "黑榜人数", 15, "人（首次触发监测督导）"])
    ws.append(["monitoring", "黑榜人次", 16, "人次 · 书面督导函"])
    ws.append(["monitoring", "头狼评选", 129, "人 · 2026 Q1"])

    # ============ M-6 培训赋能 ============
    ws = wb.create_sheet("培训赋能")
    ws.append(["section", "label", "value", "sub"])
    # 培训总量
    ws.append(["training", "能力大提升项目", 983, "场"])
    ws.append(["training", "重点培训项目", 55, "个"])
    ws.append(["training", "重点项目执行率", "58.06%", "近期启动『走进华为 2.0』"])
    # 重点培训项目
    for name, val in [
        ("广银大讲堂", 5), ("对公特训营", 2), ("个贷铁军培训", 3),
        ("办贷能力提升火线班", 4), ("校招新员工集中培训", 1),
    ]:
        ws.append(["key_projects", name, val, "期"])
    # 100% 部门
    for name in [
        "董事会办公室", "私人银行与财富管理部", "金融同业部", "金融市场部",
        "战略客户部", "消费者权益保护部", "机关纪委办公室", "安全保卫部",
    ]:
        ws.append(["dept_top", name, "100%", "完成"])
    # 底部
    for name, val in [("金融科技部", "0%"), ("内控合规部", "0%"), ("运营管理部", "14.29%")]:
        ws.append(["dept_bottom", name, val, "完成率"])
    # 资格考
    ws.append(["exam", "上岗资格考试", 35, "场次"])
    ws.append(["exam", "参考人次", 1878, "人次"])
    ws.append(["exam", "涉及岗位", 8, "大岗位"])

    out_path.parent.mkdir(exist_ok=True)
    wb.save(out_path)
    print(f"Created {out_path} ({out_path.stat().st_size} bytes)")


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/test.xlsx")
    make_test_xlsx(target)
