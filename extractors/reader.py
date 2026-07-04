"""Read .xlsx into dict[module_name, list[dict]].

v1.4: support _prev sheet (for delta calc) + optional columns delta/sub_text/metric_note.
"""
from pathlib import Path
from typing import Any

import openpyxl


def read_workbook(path: Path | str) -> dict[str, Any]:
    """读 .xlsx 全部 sheet（除 [配置] 外）→ dict[sheet_name, records] + 'prev'.

    Returns: {sheet_name: [records], 'prev': {sheet_name: [records] (from _prev sheet)}}
    后向兼容 7 列旧 Excel：missing 列默认 None。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}

    def _read_sheet(ws) -> list[dict]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        records = []
        for row in rows[1:]:
            if all(c is None or c == "" for c in row):
                continue
            record = {}
            for h, v in zip(headers, row):
                record[h] = v
            # v1.4: ensure optional fields exist (向后兼容 7 列旧 Excel)
            for opt in ("delta", "sub_text", "metric_note"):
                record.setdefault(opt, None)
            records.append(record)
        return records

    for sheet_name in wb.sheetnames:
        if sheet_name == "配置":
            continue
        ws = wb[sheet_name]
        result[sheet_name] = _read_sheet(ws)

    # v1.4: 读 _prev sheet（用于 delta 计算）
    result["prev"] = {}
    if "_prev" in wb.sheetnames:
        ws_prev = wb["_prev"]
        # _prev 结构和 _prev 内部嵌套：第一列是 sheet 名，后续是记录
        rows = list(ws_prev.iter_rows(values_only=True))
        if rows:
            # 第一行可能是 header 或直接数据。检测：如果第一行第一列是已知 sheet 名 → 直接数据
            known_sheets = {
                "M-1 组织架构", "M-2 员工情况", "M-3 人员优化",
                "M-4 干部队伍", "M-4 干部职数表", "M-5 考核薪酬", "M-6 培训赋能",
            }
            data_rows = rows[1:] if rows[0][0] in ("sheet", "表名", "模块") else rows
            buf: dict[str, list[dict]] = {}
            for r in data_rows:
                if not r or r[0] is None:
                    continue
                sheet_key = str(r[0])
                # 把整行作为 prev record；但客户端按 sheet 分组时通过 sheet_key 提取
                rec = {"_sheet": sheet_key}
                # 第一个 cell 是 sheet 名 → 跳过；记录剩余 cell 为 [分组, 名称, 数值, ...] 结构
                # 简化：期望 _prev 格式 = [sheet_name, 分组, 名称, 数值, 单位, 备注, 排序, is_total]
                cols = ["分组", "名称", "数值", "单位", "备注", "排序", "is_total"]
                for i, c in enumerate(cols):
                    if i + 1 < len(r):
                        rec[c] = r[i + 1]
                buf.setdefault(sheet_key, []).append(rec)
            result["prev"] = buf

    wb.close()
    return result


def write_prev_sheet(path: Path | str, raw_data: dict[str, list[dict]]) -> None:
    """v1.4: 把当前数据写入 _prev sheet（覆盖，下次构建用）。

    格式: [sheet_name, 分组, 名称, 数值, 单位, 备注, 排序, is_total]
    """
    wb = openpyxl.load_workbook(path)
    if "_prev" in wb.sheetnames:
        del wb["_prev"]
    ws = wb.create_sheet("_prev")
    ws.append(["sheet", "分组", "名称", "数值", "单位", "备注", "排序", "is_total"])
    for sheet_name, records in raw_data.items():
        if sheet_name == "prev":
            continue
        for r in records:
            ws.append([
                sheet_name,
                r.get("分组", ""),
                r.get("名称", ""),
                r.get("数值", ""),
                r.get("单位", ""),
                r.get("备注", ""),
                str(r.get("排序", "")),
                str(r.get("is_total", "")),
            ])
    wb.save(path)


def read_config(path: Path | str) -> dict[str, Any]:
    """读 [配置] sheet → dict[key, value]."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "配置" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["配置"]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[0] == "":
            continue
        config[str(row[0])] = row[1] if len(row) > 1 else None
    wb.close()
    return config


def clean_value(v: Any, kind: str = "auto") -> Any:
    """清洗 Excel cell 值。
    - "pct": "25.86%" → 0.2586
    - "int": "5,901" → 5901
    - "float": "3.14" → 3.14
    - empty/None → None
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return None
    if kind == "pct":
        if s.endswith("%"):
            return float(s.rstrip("%")) / 100
        return float(s)
    if kind == "int":
        return int(s.replace(",", ""))
    if kind == "float":
        return float(s.replace(",", ""))
    return s
